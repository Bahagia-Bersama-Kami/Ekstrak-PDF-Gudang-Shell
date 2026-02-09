import openpyxl
import glob
from datetime import datetime

def main():
    input_pattern = "Laporan SHELL*.xlsx"
    output_filename = "Hasil_Ekstrak_Shell_temp.xlsx"
    
    found_files = glob.glob(input_pattern)
    
    if found_files:
        source_file = found_files[0]
        print(f"--> File ditemukan: {source_file}")
        print(f"--> Sedang memproses data...")

        try:
            wb_source = openpyxl.load_workbook(source_file, data_only=True)
            
            wb_output = openpyxl.Workbook()
            default_sheet = wb_output.active
            wb_output.remove(default_sheet)

            headers = [
                "Gudang", "Tanggal", "No Inv", "No SJ", "No PO", "Tgl FP", 
                "No FP", "Byr", "Klaim/Retur", "DPP", "Rp", "JT"
            ]

            for sheet_name in wb_source.sheetnames:
                ws_source = wb_source[sheet_name]
                ws_output = wb_output.create_sheet(title=sheet_name)
                ws_output.append(headers)

                for row in ws_source.iter_rows(values_only=True):
                    val_a = row[0]
                    val_b = row[1]

                    if not val_a or not val_b:
                        continue

                    if not isinstance(val_a, str):
                        continue
                    
                    save_row = False
                    
                    if isinstance(val_b, datetime):
                        if not (val_b.year == 2001 and val_b.month == 1 and val_b.day == 1):
                            save_row = True
                    elif isinstance(val_b, str):
                        try:
                            date_obj = datetime.strptime(val_b, "%d/%m/%Y")
                            if not (date_obj.year == 2001 and date_obj.month == 1 and date_obj.day == 1):
                                save_row = True
                        except ValueError:
                            pass
                    
                    if save_row:
                        original_data = list(row[:12])
                        final_row_data = []

                        for idx, val in enumerate(original_data):
                            if idx in [1, 5, 11]:
                                if isinstance(val, datetime):
                                    final_row_data.append(val.strftime("%d/%m/%Y"))
                                else:
                                    final_row_data.append(val)
                            
                            elif idx in [9, 10]:
                                try:
                                    if val is not None:
                                        num_val = float(val)
                                        final_row_data.append(num_val)
                                    else:
                                        final_row_data.append(0)
                                except (ValueError, TypeError):
                                    final_row_data.append(val)
                            
                            else:
                                final_row_data.append(val)

                        ws_output.append(final_row_data)
                        
                        current_row = ws_output.max_row
                        cell_dpp = ws_output.cell(row=current_row, column=10)
                        cell_rp = ws_output.cell(row=current_row, column=11)
                        
                        cell_dpp.number_format = '#,##0'
                        cell_rp.number_format = '#,##0'

            wb_output.save(output_filename)
            print(f"--> Proses selesai. Data tersimpan di: {output_filename}")

        except Exception as e:
            print(f"--> Terjadi kesalahan saat memproses file: {e}")

    else:
        print(f"--> File {input_pattern} tidak ditemukan. Melewati proses ini.")

if __name__ == "__main__":
    main()