import os
import openpyxl

filename = "TEMPLATE_temp.xlsx"

if os.path.exists(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        is_changed = False
        
        for sheet in wb.worksheets:
            val_a4 = sheet.cell(row=4, column=1).value
            val_a11 = sheet.cell(row=11, column=1).value
            
            if val_a11 is not None and str(val_a4) == "JANUARI 2001":
                is_changed = True
                print(f"--> Memproses sheet: {sheet.title}")
                
                preserved_merges = []
                current_merges = list(sheet.merged_cells.ranges)
                
                for m_range in current_merges:
                    min_col, min_row, max_col, max_row = m_range.bounds
                    
                    if min_row < 4:
                        preserved_merges.append((min_col, min_row, max_col, max_row))
                    elif min_row > 10:
                        cell_val = str(sheet.cell(row=min_row, column=1).value or "")
                        if cell_val.startswith("GUDANG"):
                            preserved_merges.append((min_col, min_row - 7, max_col, max_row - 7))
                
                for m_range in current_merges:
                    sheet.unmerge_cells(str(m_range))
                
                sheet.delete_rows(4, 7)
                
                for m_bounds in preserved_merges:
                    sheet.merge_cells(
                        start_column=m_bounds[0],
                        start_row=m_bounds[1],
                        end_column=m_bounds[2],
                        end_row=m_bounds[3]
                    )
        
        if is_changed:
            wb.save(filename)
            print("--> File berhasil diperbarui")
        else:
            print("--> Tidak ada sheet yang memenuhi kriteria")
            
    except Exception as e:
        print(f"--> Terjadi kesalahan: {e}")
else:
    print(f"--> File {filename} tidak ditemukan. Proses dilewati.")