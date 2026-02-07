import os
import re
import pdfplumber
import pandas as pd
import configparser
from datetime import datetime
from openpyxl.utils import get_column_letter

def parse_date(date_str):
    if not date_str:
        return None
    clean_str = str(date_str).strip()
    clean_str = clean_str.replace('/', '.')
    try:
        return datetime.strptime(clean_str, "%d.%m.%Y")
    except ValueError:
        return None

def get_config_data():
    config = configparser.ConfigParser()
    mapping = {}
    date_filter = {"START": None, "END": None}
    
    try:
        with open('gudang.conf', 'r') as f:
            config_string = '[DEFAULT]\n' + f.read()
        config.read_string(config_string)
        
        for key, val in config['DEFAULT'].items():
            key_upper = key.upper()
            val_clean = val.strip()
            
            if key_upper == "TANGGAL_DARI":
                date_filter["START"] = parse_date(val_clean)
            elif key_upper == "TANGGAL_SAMPAI":
                date_filter["END"] = parse_date(val_clean)
            else:
                mapping[key_upper] = val_clean.upper()
                
    except Exception as e:
        print(f"--> Gagal membaca konfigurasi: {e}")
        
    return mapping, date_filter

def parse_number(num_str):
    if not num_str:
        return None
    try:
        clean = str(num_str).replace('.', '').replace(',', '.')
        if '.' in clean:
            return float(clean)
        return int(clean)
    except ValueError:
        return num_str

def parse_currency(curr_str):
    if not curr_str:
        return None
    try:
        clean = re.sub(r'[^\d,]', '', str(curr_str)) 
        clean = clean.replace('.', '').replace(',', '.') 
        
        raw_nums = re.findall(r'[\d\.,]+', str(curr_str))
        if raw_nums:
            val = raw_nums[0]
            val = val.replace('.', '') 
            val = val.replace(',', '.') 
            return float(val)
        return 0
    except Exception:
        return 0

def extract_pdf_data(filepath):
    data = {
        "Gudang": None, "Tanggal": None, "No Inv": None, "No SJ": None,
        "No PO": None, "Tgl FP": None, "No FP": None, "Byr": None,
        "Klaim/Retur": None, "DPP": None, "Rp": None, "JT": None, "Tgl Bayar": None
    }
    
    try:
        with pdfplumber.open(filepath) as pdf:
            if len(pdf.pages) > 0:
                p1 = pdf.pages[0]
                tables_p1 = p1.extract_tables()
                
                for table in tables_p1:
                    for i, row in enumerate(table):
                        row_clean = [str(c).replace('\n', ' ').strip() if c else '' for c in row]
                        
                        if any("Faktur Penjualan" in c and "Nomor" in c for c in row_clean):
                            if i + 1 < len(table):
                                val_row = table[i+1]
                                for idx, header in enumerate(row_clean):
                                    if "Faktur Penjualan" in header and "Nomor" in header:
                                        data["No Inv"] = parse_number(val_row[idx])
                                    if "Tanggal" in header:
                                        dt = parse_date(val_row[idx])
                                        data["Tanggal"] = dt
                                        data["Tgl FP"] = dt
                                    if "Due Date" in header:
                                        data["JT"] = parse_date(val_row[idx])
                                        
                        if any("Total Amount (Incl all taxes)" in c for c in row_clean):
                            row_str = " ".join(row_clean)
                            match = re.search(r"IDR\s*([\d\.,]+)", row_str)
                            if match:
                                data["Rp"] = parse_currency(match.group(1))
                            else:
                                for val in row_clean:
                                    if re.match(r'^[\d\.,]+$', val) and len(val) > 3:
                                        data["Rp"] = parse_currency(val)
                                        
                        if any("Total Amount (Excl Tax Amount)" in c for c in row_clean):
                            row_str = " ".join(row_clean)
                            match = re.search(r"IDR\s*([\d\.,]+)", row_str)
                            if match:
                                data["DPP"] = parse_currency(match.group(1))
                            else:
                                for val in row_clean:
                                    if re.match(r'^[\d\.,]+$', val) and len(val) > 3:
                                        data["DPP"] = parse_currency(val)

            if len(pdf.pages) > 1:
                p2 = pdf.pages[1]
                tables_p2 = p2.extract_tables()
                
                for table in tables_p2:
                    for row in table:
                        row_clean = [str(c).replace('\n', ' ').strip() if c else '' for c in row]
                        
                        for idx, cell in enumerate(row_clean):
                            if "Your Reference" in cell:
                                if idx + 1 < len(row_clean):
                                    val = row_clean[idx+1]
                                    if val:
                                        data["Gudang"] = val.split()[0].strip().upper()
                            
                            if "No.Pemesanan" in cell:
                                if idx + 1 < len(row_clean):
                                    data["No PO"] = parse_number(row_clean[idx+1])

                            if cell == "Nomor":
                                if idx + 1 < len(row_clean):
                                    val = row_clean[idx+1]
                                    if val and "/" in val:
                                        clean_sj = val.split('/')[0].strip()
                                        data["No SJ"] = parse_number(clean_sj)

                text_p2 = p2.extract_text()
                if text_p2:
                    if not data["Gudang"]:
                        match = re.search(r"Your Reference\s+([A-Za-z0-9]+)", text_p2)
                        if match:
                            data["Gudang"] = match.group(1).upper()
                    
                    if not data["No PO"]:
                        match = re.search(r"No\.Pemesanan\s+([0-9]+)", text_p2)
                        if match:
                            data["No PO"] = parse_number(match.group(1))

                    if not data["No SJ"]:
                        match = re.search(r"Nomor\s+([0-9]+)\s*\/", text_p2)
                        if match:
                            data["No SJ"] = parse_number(match.group(1))

    except Exception as e:
        print(f"--> Error pada file {filepath}: {e}")

    return data

def main():
    mapping, date_filter = get_config_data()
    files = [f for f in os.listdir('.') if f.lower().endswith('.pdf')]
    
    if not files:
        print("--> Tidak ada file PDF ditemukan.")
        return

    print(f"--> Menemukan {len(files)} file PDF.")
    all_data = []

    for f in files:
        print(f"--> Memproses: {f}")
        res = extract_pdf_data(f)
        
        raw_gudang = res["Gudang"]
        if raw_gudang in mapping:
            sheet_code = mapping[raw_gudang]
            res["Gudang"] = sheet_code
            sheet_name = sheet_code
        else:
            sheet_name = "DATA_LAIN"
            
        res["_SHEET"] = sheet_name
        all_data.append(res)

    if not all_data:
        print("--> Tidak ada data berhasil diekstrak.")
        return

    df = pd.DataFrame(all_data)
    cols = ["Gudang", "Tanggal", "No Inv", "No SJ", "No PO", "Tgl FP", "No FP", 
            "Byr", "Klaim/Retur", "DPP", "Rp", "JT", "Tgl Bayar", "_SHEET"]
    df = df[cols]
    
    if "Tanggal" in df.columns and date_filter["START"] and date_filter["END"]:
        start_d = date_filter["START"]
        end_d = date_filter["END"]
        print(f"--> Memfilter data dari {start_d.strftime('%d/%m/%Y')} s/d {end_d.strftime('%d/%m/%Y')}")
        
        df = df.dropna(subset=['Tanggal'])
        
        mask = (df['Tanggal'] >= start_d) & (df['Tanggal'] <= end_d)
        df = df[mask]
        
        if df.empty:
            print("--> Semua data terhapus setelah difilter tanggal.")
            return

    if "Tanggal" in df.columns:
        df.sort_values(by="Tanggal", ascending=True, inplace=True)

    output_file = "Hasil_Ekstrak_temp.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        unique_sheets = df["_SHEET"].unique()
        for sheet in unique_sheets:
            dsub = df[df["_SHEET"] == sheet].drop(columns=["_SHEET"])
            dsub.to_excel(writer, sheet_name=sheet, index=False)
            
            worksheet = writer.sheets[sheet]
            
            for row in worksheet.iter_rows(min_row=2):
            	
                cell_b = row[1] 
                cell_b.number_format = 'DD/MM/YYYY'
                
                cell_f = row[5]
                cell_f.number_format = 'DD/MM/YYYY'
                
                cell_j = row[9]
                cell_j.number_format = '#,##0'
                
                cell_k = row[10]
                cell_k.number_format = '#,##0'
                
                cell_l = row[11]
                cell_l.number_format = 'DD/MM/YYYY'

            for column in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        val_len = len(str(cell.value))
                        if val_len > max_len:
                            max_len = val_len
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = max_len + 4

    print(f"--> Selesai. Data disimpan di {output_file}")

if __name__ == "__main__":
    main()