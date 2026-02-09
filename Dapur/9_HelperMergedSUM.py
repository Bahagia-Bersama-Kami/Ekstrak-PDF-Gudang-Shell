import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

file_name = 'TEMPLATE_temp.xlsx'

if os.path.exists(file_name):
    print(f"--> File {file_name} ditemukan. Memulai proses.")
    
    wb = load_workbook(file_name)
    
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for sheet in wb.worksheets:
        print(f"--> Memeriksa sheet: {sheet.title}")
        
        gudang_rows = []
        for row in sheet.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and isinstance(cell.value, str) and "GUDANG" in cell.value:
                gudang_rows.append(cell.row)
        
        if not gudang_rows:
            print(f"--> Tidak ada marker GUDANG di sheet {sheet.title}, skip.")
            continue
            
        current_data_start = gudang_rows[0] + 2
        
        for i in range(1, len(gudang_rows)):
            trigger_row = gudang_rows[i]
            target_row = trigger_row - 4 
            
            if target_row > current_data_start:
                sum_end = target_row - 1
                
                sheet.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=10)
                
                for c in range(1, 11): 
                    sheet.cell(row=target_row, column=c).border = border_style
                
                target_cell = sheet.cell(row=target_row, column=11)
                target_cell.value = f"=SUM(K{current_data_start}:K{sum_end})"
                
                print(f"--> [Tabel Tengah] Merger Baris {target_row}. Range SUM: K{current_data_start}:K{sum_end}")
                
            current_data_start = trigger_row + 2
            
        last_data_row = 0
        for r in range(sheet.max_row, current_data_start - 1, -1):
            val = sheet.cell(row=r, column=1).value
            
            if val and isinstance(val, str) and "GUDANG" not in val:
                last_data_row = r
                break
                
        if last_data_row >= current_data_start:
            target_row_last = last_data_row + 1 
            
            sheet.merge_cells(start_row=target_row_last, start_column=1, end_row=target_row_last, end_column=10)
            for c in range(1, 11): 
                sheet.cell(row=target_row_last, column=c).border = border_style
            
            target_cell_last = sheet.cell(row=target_row_last, column=11)
            target_cell_last.value = f"=SUM(K{current_data_start}:K{last_data_row})"
            
            print(f"--> [Tabel Terakhir] Merger Baris {target_row_last}. Range SUM: K{current_data_start}:K{last_data_row}")

    wb.save(file_name)
    print("--> Selesai. File telah disimpan dan diperbarui.")

else:
    print(f"--> File {file_name} tidak ditemukan. Proses dilewati.")