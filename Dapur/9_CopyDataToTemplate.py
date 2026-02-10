import openpyxl
import os
import shutil
from datetime import datetime
from copy import copy
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def get_indonesian_month(month_idx):
    months = {
        1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
        5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
        9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
    }
    return months.get(month_idx, "")

def get_thin_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def apply_header_structure(sheet, start_row):
    for col in range(1, 8):
        sheet.merge_cells(start_row=start_row, start_column=col, end_row=start_row+1, end_column=col)
        cell = sheet.cell(row=start_row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.merge_cells(start_row=start_row, start_column=8, end_row=start_row, end_column=9)
    cell_debet = sheet.cell(row=start_row, column=8)
    cell_debet.value = "Debet"
    cell_debet.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=start_row, start_column=10, end_row=start_row+1, end_column=10)
    cell_dpp = sheet.cell(row=start_row, column=10)
    cell_dpp.value = "DPP"
    cell_dpp.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=13)
    cell_kredit = sheet.cell(row=start_row, column=11)
    cell_kredit.value = "Kredit"
    cell_kredit.alignment = Alignment(horizontal='center', vertical='center')

def main():
    source_file = "Hasil_Ekstrak_temp.xlsx"
    template_file = "TEMPLATE.xlsx"
    target_file = "TEMPLATE_temp.xlsx"
    
    shutil.copy(template_file, target_file)
    print(f"--> Menggunakan file target: {target_file}")

    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    wb_target = openpyxl.load_workbook(target_file)
    wb_template = openpyxl.load_workbook(template_file) 

    print(f"--> Membaca data dari {source_file}")

    for sheet_name in wb_source.sheetnames:
        if sheet_name not in wb_target.sheetnames or sheet_name not in wb_template.sheetnames:
            continue

        print(f"--> Memproses Sheet: {sheet_name}")
        
        ws_source = wb_source[sheet_name]
        ws_target = wb_target[sheet_name]
        ws_tmpl = wb_template[sheet_name]

        data_by_month = {}
        max_col_source = ws_source.max_column 
        max_col_visual = 14 
        
        for row in ws_source.iter_rows(min_row=2, values_only=True):
            if not row[1]: 
                continue
            
            date_val = row[1]
            date_obj = None
            if isinstance(date_val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        date_obj = datetime.strptime(date_val, fmt)
                        break
                    except ValueError:
                        pass
            elif isinstance(date_val, datetime):
                date_obj = date_val
            
            if not date_obj:
                continue

            month_key = (date_obj.year, date_obj.month)
            if month_key not in data_by_month:
                data_by_month[month_key] = []
            data_by_month[month_key].append(row)

        sorted_months = sorted(data_by_month.keys())
        
        for year, month in sorted_months:
            month_name = get_indonesian_month(month)
            header_text = f"{month_name} {year}"
            new_data_rows = data_by_month[(year, month)]

            found_header_row = -1
            for r in range(1, ws_target.max_row + 1):
                cell_val = ws_target.cell(row=r, column=1).value
                if cell_val and str(cell_val).strip().upper() == header_text:
                    found_header_row = r
                    break
                    
            if found_header_row != -1:
                print(f"--> Data {header_text} ditemukan. Menambahkan data...")
                
                data_start_row = found_header_row + 3
                sum_row = -1
                current_r = data_start_row
                
                while current_r <= ws_target.max_row + 500:
                    cell_k = ws_target.cell(row=current_r, column=11)
                    is_merged_a = False
                    for merged in ws_target.merged_cells.ranges:
                         if ws_target.cell(row=current_r, column=1).coordinate in merged:
                             is_merged_a = True
                             break
                    if (cell_k.value and str(cell_k.value).strip().startswith("=")) or is_merged_a:
                        sum_row = current_r
                        break
                    current_r += 1
                
                if sum_row == -1:
                    continue

                existing_sigs = set()
                for r in range(data_start_row, sum_row):
                    vals = []
                    for c in range(1, max_col_source + 1):
                        vals.append(ws_target.cell(row=r, column=c).value)
                    sig = tuple(str(v) if v is not None else "" for v in vals)
                    existing_sigs.add(sig)

                rows_to_insert = []
                for row_data in new_data_rows:
                    vals = list(row_data)[:max_col_source]
                    sig = tuple(str(v) if v is not None else "" for v in vals)
                    if sig not in existing_sigs:
                        rows_to_insert.append(row_data)

                if rows_to_insert:
                    print(f"--> Menyisipkan {len(rows_to_insert)} baris.")
                    ws_target.insert_rows(sum_row, amount=len(rows_to_insert))
                    
                    for idx, data_row in enumerate(rows_to_insert):
                        target_r = sum_row + idx
                        try:
                            ws_target.unmerge_cells(start_row=target_r, start_column=1, end_row=target_r, end_column=10)
                        except:
                            pass

                        for col_idx in range(1, max_col_visual + 1):
                            new_cell = ws_target.cell(row=target_r, column=col_idx)
                            if col_idx <= len(data_row):
                                new_cell.value = data_row[col_idx-1]
                                
                            tmpl_cell = ws_tmpl.cell(row=7, column=col_idx)
                            copy_cell_style(tmpl_cell, new_cell)
                            new_cell.border = get_thin_border()
                            
                    new_sum_row = sum_row + len(rows_to_insert)
                    cell_sum = ws_target.cell(row=new_sum_row, column=11)
                    
                    style_tmpl_sum = ws_tmpl.cell(row=7, column=11)
                    copy_cell_style(style_tmpl_sum, cell_sum) 
                    cell_sum.border = get_thin_border()
                    
                    cell_sum.value = f"=SUM(K{data_start_row}:K{new_sum_row-1})"
                    cell_sum.font = Font(color="FF0000", bold=True)
                    
                    try:
                        ws_target.merge_cells(start_row=new_sum_row, start_column=1, end_row=new_sum_row, end_column=10)
                    except:
                        pass
                        
            else:
                print(f"--> Membuat blok baru untuk {header_text}")
                
                last_row = ws_target.max_row
                start_block_row = last_row + 3
                
                ws_target.cell(row=start_block_row, column=1).value = header_text
                for c in range(1, max_col_visual + 5): 
                    copy_cell_style(ws_tmpl.cell(row=4, column=c), ws_target.cell(row=start_block_row, column=c))
                    
                for r_offset in [1, 2]:
                    target_r = start_block_row + r_offset
                    tmpl_r = 4 + r_offset 
                    for c in range(1, max_col_visual + 1): 
                        tmpl_cell = ws_tmpl.cell(row=tmpl_r, column=c)
                        target_cell = ws_target.cell(row=target_r, column=c)
                        target_cell.value = tmpl_cell.value
                        copy_cell_style(tmpl_cell, target_cell)

                apply_header_structure(ws_target, start_block_row + 1)
                
                current_data_row = start_block_row + 3
                for row_data in new_data_rows:
                    for col_idx in range(1, max_col_visual + 1):
                        target_cell = ws_target.cell(row=current_data_row, column=col_idx)
                        if col_idx <= len(row_data):
                            target_cell.value = row_data[col_idx-1]
                        
                        style_source = ws_tmpl.cell(row=7, column=col_idx)
                        copy_cell_style(style_source, target_cell)
                        target_cell.border = get_thin_border()
                    
                    current_data_row += 1
                    
                sum_row_pos = current_data_row
                ws_target.merge_cells(start_row=sum_row_pos, start_column=1, end_row=sum_row_pos, end_column=10)
                
                data_start_idx = start_block_row + 3
                data_end_idx = sum_row_pos - 1
                
                for c in range(1, max_col_visual + 1):
                    target_c = ws_target.cell(row=sum_row_pos, column=c)
                    style_c = ws_tmpl.cell(row=7, column=c)
                    copy_cell_style(style_c, target_c)
                    target_c.border = get_thin_border()
                    
                cell_sum = ws_target.cell(row=sum_row_pos, column=11)
                cell_sum.value = f"=SUM(K{data_start_idx}:K{data_end_idx})"
                cell_sum.font = Font(color="FF0000", bold=True)

    wb_target.save(target_file)
    print(f"--> Proses selesai. File tersimpan: {target_file}")

if __name__ == "__main__":
    main()
