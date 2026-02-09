import openpyxl
import os
from datetime import datetime

def parse_date(date_val):
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, "%d/%m/%Y")
        except ValueError:
            pass
    return datetime.min

def format_date_value(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str):
        try:
            if " " in value:
                dt = datetime.strptime(value.split()[0], "%Y-%m-%d")
                return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return value

def main():
    source_file = "Hasil_Ekstrak_Shell_temp.xlsx"
    target_file = "Hasil_Ekstrak_temp.xlsx"

    if not os.path.exists(source_file):
        print(f"--> File {source_file} tidak ditemukan. Proses ini dilewati.")
        return

    if not os.path.exists(target_file):
        print(f"--> File target {target_file} tidak ditemukan. Harap sediakan file target.")
        return

    print(f"--> Memproses data dari {source_file} ke {target_file}")

    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    wb_target = openpyxl.load_workbook(target_file)
    
    date_columns_indices = [1, 5, 11]

    for sheet_name in wb_source.sheetnames:
        if sheet_name not in wb_target.sheetnames:
            continue

        ws_source = wb_source[sheet_name]
        ws_target = wb_target[sheet_name]

        print(f"--> Mengolah sheet: {sheet_name}")

        new_data = []
        for row in ws_source.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                new_data.append(list(row))

        existing_data = []
        header = []
        for i, row in enumerate(ws_target.iter_rows(values_only=True)):
            if i == 0:
                header = list(row)
            else:
                if any(cell is not None for cell in row):
                    existing_data.append(list(row))

        combined_data = existing_data + new_data
        
        # Hapus Duplikat
        unique_data = []
        seen = set()
        
        for row in combined_data:
            row_tuple = tuple(row)
            if row_tuple not in seen:
                seen.add(row_tuple)
                unique_data.append(row)
                
        sort_col_idx = 1 
        
        unique_data.sort(key=lambda x: parse_date(x[sort_col_idx]))
        
        if ws_target.max_row >= 2:
            ws_target.delete_rows(2, amount=ws_target.max_row)
            
        for row_data in unique_data:
            final_row = list(row_data)
            
            for idx in date_columns_indices:
                if idx < len(final_row):
                    final_row[idx] = format_date_value(final_row[idx])
            
            ws_target.append(final_row)

    wb_target.save(target_file)
    print(f"--> Proses selesai. Data tersimpan, terurut, dan diformat di: {target_file}")

if __name__ == "__main__":
    main()